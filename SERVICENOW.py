import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime


def combinar_arquivos():
    pasta = r"Z:\#BASES DE APOIO - GERENCIAL\01 - BASES\PROJETOS\SERVICE NOW"
    dfs = []

    for filename in os.listdir(pasta):
        if filename.endswith(".xlsx") and not filename.startswith('~$'):
            arquivo = os.path.join(pasta, filename)
            
            try:
                df = pd.read_excel(arquivo, sheet_name="Page 1")
                dfs.append(df)
                print(f"Arquivo '{filename}' lido com sucesso.")
            
            except Exception as e:
                print(f"Erro ao ler o arquivo '{filename}': {e}")

    if dfs:
        df_final = pd.concat(dfs, ignore_index=True)
        default_font = Font(name='Calibri', size=12)
        wb = Workbook()
        ws = wb.active
        
        for r in dataframe_to_rows(df_final, index=False, header=True):
            ws.append(r)
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = default_font
        
        output_directory = r"C:\Users\2160011883.VIA\OneDrive - Via Varejo S.A\GLAUBER S-ONE DRIVE\PROJETOS\PYTHON\Analise de dados\ETL\Pbi\Estudo de verbas"
        output_file = os.path.join(output_directory, "f_Servicenow_combinado.xlsx")
        
        if os.path.exists(output_file):
            os.remove(output_file)
            print(f"Arquivo existente '{output_file}' removido.")
        
        wb.save(output_file)
        print(f"Arquivo combinado salvo em: {output_file}")
    else:
        print("Nenhum arquivo válido encontrado para combinar.")
        

    # df = pd.read_excel(output_file)
    # colunas_desejadas = [
    #     "Item","Número OM"
    # ]
    # colunas_presentes = [col for col in colunas_desejadas if col in df.columns]
    # df_reduzido = df[colunas_presentes]

    
    # output_path = r"C:\Users\2160011883\OneDrive - Via Varejo S.A\GLAUBER S-ONE DRIVE\PROJETOS\PYTHON\Analise de dados\ETL\Pbi\Estudo de verbas"
    # output_file = os.path.join(output_path, "f_SN.xlsx")

    # if os.path.exists(output_file):
    #     os.remove(output_file)
    #     print(f"Arquivo existente '{output_file}' removido.")

    # df_reduzido.to_excel(output_file, index=False)
    # print(f"DataFrame processado salvo com sucesso em '{output_file}'")


combinar_arquivos()    
